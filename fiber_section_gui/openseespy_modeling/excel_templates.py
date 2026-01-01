# -*- coding: utf-8 -*-
"""
Excel模板生成模块
用于创建节点和单元批量导入的Excel模板
"""

import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from PyQt5.QtCore import QObject, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl.utils import get_column_letter

from .model_settings import ModelSettings
from .node_manager import NodeManager
from .element_manager import ElementManager


class ExcelTemplates(QObject):
    """Excel模板生成器"""
    
    # 信号定义
    template_created = pyqtSignal(str, str)  # 模板创建信号 (文件路径, 模板类型)
    template_error = pyqtSignal(str)  # 模板创建错误信号
    
    def __init__(self, model_settings: ModelSettings, node_manager: NodeManager, 
                 element_manager: ElementManager):
        super().__init__()
        self.model_settings = model_settings
        self.node_manager = node_manager
        self.element_manager = element_manager
        
    def create_node_template(self, file_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        创建节点批量导入Excel模板
        
        Args:
            file_path: 文件路径，如果为None则弹出保存对话框
            
        Returns:
            Tuple[bool, str]: (是否成功, 错误信息)
        """
        try:
            if file_path is None:
                # 在测试环境中，如果没有QApplication，使用默认路径
                try:
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存节点批量导入模板",
                        "nodes_template.xlsx",
                        "Excel Files (*.xlsx);;All Files (*)"
                    )
                except Exception:
                    # 如果GUI不可用，使用默认文件名
                    file_path = "nodes_template.xlsx"
                
            if not file_path:
                return False, "用户取消了文件保存"
                
            # 创建Excel写入器
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                
                # 创建说明页
                self._create_node_instruction_sheet(writer)
                
                # 创建节点数据模板页
                self._create_node_data_sheet(writer)
                
                # 创建示例数据页
                self._create_node_example_sheet(writer)
                
            self.template_created.emit(file_path, "节点")
            return True, f"成功创建节点模板: {file_path}"
            
        except Exception as e:
            error_msg = f"创建节点模板失败: {str(e)}"
            self.template_error.emit(error_msg)
            return False, error_msg
            
    def create_element_template(self, file_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        创建单元批量导入Excel模板
        
        Args:
            file_path: 文件路径，如果为None则弹出保存对话框
            
        Returns:
            Tuple[bool, str]: (是否成功, 错误信息)
        """
        try:
            if file_path is None:
                # 在测试环境中，如果没有QApplication，使用默认路径
                try:
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存单元批量导入模板",
                        "elements_template.xlsx",
                        "Excel Files (*.xlsx);;All Files (*)"
                    )
                except Exception:
                    # 如果GUI不可用，使用默认文件名
                    file_path = "elements_template.xlsx"
                
            if not file_path:
                return False, "用户取消了文件保存"
                
            # 创建Excel写入器
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                
                # 创建说明页
                self._create_element_instruction_sheet(writer)
                
                # 为每种单元类型创建单独的页
                element_types = self.element_manager.get_element_types()
                for elem_type in element_types:
                    self._create_element_data_sheet(writer, elem_type)
                    
                # 创建示例数据页
                self._create_element_example_sheet(writer)
                
            self.template_created.emit(file_path, "单元")
            return True, f"成功创建单元模板: {file_path}"
            
        except Exception as e:
            error_msg = f"创建单元模板失败: {str(e)}"
            self.template_error.emit(error_msg)
            return False, error_msg
            
    def _create_node_instruction_sheet(self, writer):
        """创建节点模板说明页"""
        instructions = [
            ["节点批量导入说明", ""],
            ["", ""],
            ["1. 基本要求", ""],
            ["- 节点ID必须是唯一的正整数", ""],
            ["- 坐标值根据模型维度确定（2D: x,y；3D: x,y,z）", ""],
            ["- 质量数据长度为自由度数量，默认6个分量", ""],
            ["", ""],
            ["2. 列说明", ""],
            ["id", "节点ID（必需）"],
            ["x", "X坐标（必需）"],
            ["y", "Y坐标（必需）"],
            ["z", "Z坐标（仅3D模型必需，2D模型可为0）"],
            ["mass", "质量数据，用逗号分隔（如：0,0,0,0,0,0）"],
            ["name", "节点名称（可选）"],
            ["", ""],
            ["3. 示例数据", ""],
            ["请参考'示例数据'工作表"],
            ["", ""],
            ["4. 注意事项", ""],
            ["- 不要删除或重命名列", ""],
            ["- 保持数据格式一致", ""],
            ["- ID冲突会导致导入失败", ""],
            ["- 可以从Excel复制数据到GUI中粘贴"]
        ]
        
        df = pd.DataFrame(instructions, columns=['说明', '备注'])
        df.to_excel(writer, sheet_name='使用说明', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['使用说明']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 50
        
    def _create_node_data_sheet(self, writer):
        """创建节点数据模板页"""
        # 根据模型维度确定列
        if self.model_settings.ndm == 3:
            columns = ['id', 'x', 'y', 'z', 'mass', 'name']
            default_mass = ','.join(['0.0'] * self.model_settings.ndf)
            sample_data = [
                [1, 0.0, 0.0, 0.0, default_mass, 'Node_1'],
                [2, 1.0, 0.0, 0.0, default_mass, 'Node_2'],
                [3, 0.0, 1.0, 0.0, default_mass, 'Node_3']
            ]
        else:  # 2D
            columns = ['id', 'x', 'y', 'mass', 'name']
            default_mass = ','.join(['0.0'] * self.model_settings.ndf)
            sample_data = [
                [1, 0.0, 0.0, default_mass, 'Node_1'],
                [2, 1.0, 0.0, default_mass, 'Node_2'],
                [3, 0.0, 1.0, default_mass, 'Node_3']
            ]
            
        df = pd.DataFrame(sample_data, columns=columns)
        df.to_excel(writer, sheet_name='节点数据', index=False)
        
        # 设置列宽 - 使用openpyxl的列字母
        worksheet = writer.sheets['节点数据']
        for i, col in enumerate(columns):
            col_letter = get_column_letter(i + 1)  # i从0开始，但get_column_letter从1开始
            worksheet.column_dimensions[col_letter].width = 15
            
    def _create_node_example_sheet(self, writer):
        """创建节点示例数据页"""
        if self.model_settings.ndm == 3:
            columns = ['id', 'x', 'y', 'z', 'mass', 'name']
            example_data = [
                [1, 0.0, 0.0, 0.0, '0,0,0,0,0,0', 'Origin'],
                [2, 2.0, 0.0, 0.0, '0,0,0,0,0,0', 'X_axis'],
                [3, 0.0, 2.0, 0.0, '0,0,0,0,0,0', 'Y_axis'],
                [4, 0.0, 0.0, 2.0, '0,0,0,0,0,0', 'Z_axis'],
                [5, 1.0, 1.0, 1.0, '1.0,1.0,1.0,0.1,0.1,0.1', 'Center']
            ]
        else:
            columns = ['id', 'x', 'y', 'mass', 'name']
            example_data = [
                [1, 0.0, 0.0, '0,0,0', 'Origin'],
                [2, 2.0, 0.0, '0,0,0', 'X_axis'],
                [3, 0.0, 2.0, '0,0,0', 'Y_axis'],
                [4, 1.0, 1.0, '1.0,1.0,0.1', 'Center']
            ]
            
        df = pd.DataFrame(example_data, columns=columns)
        df.to_excel(writer, sheet_name='示例数据', index=False)
        
        # 设置列宽 - 使用openpyxl的列字母
        worksheet = writer.sheets['示例数据']
        for i, col in enumerate(columns):
            col_letter = get_column_letter(i + 1)  # i从0开始，但get_column_letter从1开始
            worksheet.column_dimensions[col_letter].width = 15
            
    def _create_element_instruction_sheet(self, writer):
        """创建单元模板说明页"""
        instructions = [
            ["单元批量导入说明", ""],
            ["", ""],
            ["1. 基本要求", ""],
            ["- 单元ID必须是唯一的正整数", ""],
            ["- 节点ID必须已存在", ""],
            ["- 不同单元类型需要不同参数", ""],
            ["", ""],
            ["2. 单元类型说明", ""],
            ["ZeroLength", "零长度单元（2节点）"],
            ["TwoNodeLink", "双节点连接单元（2节点）"],
            ["Truss", "桁架单元（2节点）"],
            ["ElasticBeamColumn", "弹性梁柱单元（2节点）"],
            ["DispBeamColumn", "位移梁柱单元（2节点）"],
            ["ForceBeamColumn", "力梁柱单元（2节点）"],
            ["", ""],
            ["3. 数据格式", ""],
            ["- 每种单元类型在单独的工作表中", ""],
            ["- 必须包含列：id, node1, node2", ""],
            ["- 其他列根据单元类型确定", ""],
            ["", ""],
            ["4. 注意事项", ""],
            ["- 不要删除或重命名必需列", ""],
            ["- 节点ID必须存在且有效", ""],
            ["- 参数值必须在合理范围内"]
        ]
        
        df = pd.DataFrame(instructions, columns=['说明', '备注'])
        df.to_excel(writer, sheet_name='使用说明', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['使用说明']
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 50
        
    def _create_element_data_sheet(self, writer, element_type: str):
        """创建特定单元类型的数据模板页"""
        sheet_name = f"{element_type}_数据"
        
        # 根据单元类型定义列
        if element_type == 'ZeroLength':
            columns = ['id', 'node1', 'node2', 'mat_tags', 'dirs', 'do_rayleigh', 'r_flag']
            sample_data = [
                [1, 1, 2, '1,2', '1,2', 'False', 0],
                [2, 2, 3, '3', '1', 'True', 1]
            ]
        elif element_type == 'Truss':
            columns = ['id', 'node1', 'node2', 'A', 'mat_tag', 'rho', 'c_mass', 'do_rayleigh']
            sample_data = [
                [1, 1, 2, 0.01, 1, 7850.0, 'False', 'False'],
                [2, 2, 3, 0.015, 2, 7850.0, 'True', 'False']
            ]
        elif element_type == 'ElasticBeamColumn':
            columns = ['id', 'node1', 'node2', 'Area', 'E_mod', 'Iz', 'transf_tag', 'mass', 'c_mass']
            sample_data = [
                [1, 1, 2, 0.01, 200000, 8.33e-6, 1, 0.0, 'False'],
                [2, 2, 3, 0.015, 200000, 1.25e-5, 2, 0.0, 'False']
            ]
        elif element_type == 'TwoNodeLink':
            columns = ['id', 'node1', 'node2', 'mat_tags', 'dirs', 'p_delta', 'shear_dist', 'do_rayleigh', 'mass']
            sample_data = [
                [1, 1, 2, '1,2', '1,2', '', '', 'False', 0.0],
                [2, 2, 3, '3', '1', '0.5', '0.5', 'True', 1.0]
            ]
        elif element_type == 'DispBeamColumn':
            columns = ['id', 'node1', 'node2', 'transf_tag', 'integration_tag', 'c_mass', 'mass']
            sample_data = [
                [1, 1, 2, 1, 1, 'True', 0.0],
                [2, 2, 3, 2, 2, 'False', 1.0]
            ]
        elif element_type == 'ForceBeamColumn':
            columns = ['id', 'node1', 'node2', 'transf_tag', 'integration_tag', 'max_iter', 'tol', 'mass']
            sample_data = [
                [1, 1, 2, 1, 1, 10, 1e-12, 0.0],
                [2, 2, 3, 2, 2, 15, 1e-10, 1.0]
            ]
        else:
            # 通用格式
            columns = ['id', 'node1', 'node2']
            sample_data = [[1, 1, 2]]
            
        df = pd.DataFrame(sample_data, columns=columns)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置列宽
        worksheet = writer.sheets[sheet_name]
        for i, col in enumerate(columns):
            col_letter = get_column_letter(i + 1)  # i从0开始，但get_column_letter从1开始
            worksheet.column_dimensions[col_letter].width = 15
            
    def _create_element_example_sheet(self, writer):
        """创建单元示例数据页"""
        examples = [
            ["单元类型", "示例说明"],
            ["ZeroLength", "零长度单元常用于连接不同节点，如基础连接"],
            ["TwoNodeLink", "双节点连接单元可用于模拟隔震支座等"],
            ["Truss", "桁架单元用于模拟轴向受力构件"],
            ["ElasticBeamColumn", "弹性梁柱单元用于模拟弹性梁柱"],
            ["DispBeamColumn", "位移梁柱单元用于非线性分析"],
            ["ForceBeamColumn", "力梁柱单元用于高精度的非线性分析"],
            ["", ""],
            ["参数说明", ""],
            ["mat_tags", "材料标签列表，用逗号分隔"],
            ["dirs", "方向标签列表，用逗号分隔"],
            ["A", "截面积"],
            ["mat_tag", "材料标签"],
            ["Area", "截面积"],
            ["E_mod", "弹性模量"],
            ["Iz", "惯性矩"],
            ["transf_tag", "坐标变换标签"],
            ["integration_tag", "积分点标签"]
        ]
        
        df = pd.DataFrame(examples[1:], columns=examples[0])
        df.to_excel(writer, sheet_name='示例说明', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['示例说明']
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 50
        
    def create_all_templates(self, directory: Optional[str] = None) -> Tuple[bool, str, List[str]]:
        """
        创建所有模板
        
        Args:
            directory: 保存目录，如果为None则弹出选择对话框
            
        Returns:
            Tuple[bool, str, List[str]]: (是否成功, 错误信息, 创建的文件列表)
        """
        try:
            if directory is None:
                directory = QFileDialog.getExistingDirectory(
                    None,
                    "选择模板保存目录"
                )
                
            if not directory:
                return False, "用户取消了目录选择", []
                
            created_files = []
            
            # 创建节点模板
            node_file = os.path.join(directory, "nodes_template.xlsx")
            success, msg = self.create_node_template(node_file)
            if success:
                created_files.append(node_file)
            else:
                return False, msg, []
                
            # 创建单元模板
            element_file = os.path.join(directory, "elements_template.xlsx")
            success, msg = self.create_element_template(element_file)
            if success:
                created_files.append(element_file)
            else:
                return False, msg, []
                
            return True, f"成功创建所有模板", created_files
            
        except Exception as e:
            return False, f"创建模板时发生错误: {str(e)}", []
            
    def open_template_in_excel(self, template_type: str, file_path: str) -> bool:
        """
        在Excel中打开模板
        
        Args:
            template_type: 模板类型 ('node' 或 'element')
            file_path: 模板文件路径
            
        Returns:
            bool: 是否成功打开
        """
        try:
            import subprocess
            import platform
            
            # 根据操作系统打开文件
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', file_path])
            else:  # Linux
                subprocess.call(['xdg-open', file_path])
                
            return True
            
        except Exception:
            return False